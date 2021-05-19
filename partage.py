from openpyxl import load_workbook
import datetime
from collections import defaultdict
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import locale

FIRST_ROW = 5
COLUMNS = {
    "name": 1,
    "year_frequency": 2,
    "time": 3,
    "penibility": 5,
    "attributed": 7,
    "shift": 8,
    "hide": 9,
}
WEEKS_IN_YEAR = 52
MAX_ROWS = 100
ONE_WEEK = datetime.timedelta(days=7)
ONE_DAY = datetime.timedelta(days=1)
FIRST_DAY = datetime.datetime(2021, 4, 14) - ONE_WEEK
LAST_DAY = datetime.datetime(2021, 12, 31)
# lookup in `attributed` lowercase string, participant name
PARTICIPANTS = (
    ('x', 'Maxime'),
    ('o', 'Morgane'),
)
WORBOOK_PATH = "2021 - Partage du ménage Chouchou et Chouchou.xlsx"
NAME_STRING = "Semaine du \n\nau"
OUTPUT_SHEET = "Détails"
OUTPUT_FIRST_ROW = 2
TASK_FORMAT = "☐ {name} ({time:.0f}min)"
TASK_FONT = Font(size=12, color='000000')  # black
WEEK_FONT = Font(size=14, color='808080')  # grey
PARTICIPANT_FONT = Font(size=20, color='000000')  # black
THIN = Side(border_style="thin", color="000000")
MEDIUM = Side(border_style="medium", color="000000")
OUTER_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER_ALIGNMENT = Alignment(wrap_text=True, horizontal='center', vertical='center')
TODO_PARTICIPANT = "à faire {}"
TASK_CELL_ALIGNMENT = Alignment(horizontal='general', vertical='center', wrap_text=True)
TASK_ROW_HEIGHT = 280
TOP_ROW_HEIGHT = 40
TASK_COLUMN_WIDTH = 80
WEEK_COLUMN_WIDTH = 30


def get_row(ws, row):
    to_return = {
        key: ws.cell(row, col).value
        for key, col in COLUMNS.items()
    }
    if not to_return["year_frequency"]:
        return None
    to_return["shift"] = to_return["shift"] or 0

    return to_return


def serialize_date(date, year=False):
    if year:
        return date.strftime("%d %b %Y")
    return date.strftime("%d %b")


def rest_after_division(to_divide, divide_by):
    return to_divide - int(to_divide / divide_by) * divide_by


def should_do_task(task_year_frequency, week_number, task_shift):
    return rest_after_division(
        week_number, WEEKS_IN_YEAR / task_year_frequency
    ) < 1


def get_week_number(day):
    return day.isocalendar()[1]


def format_tasks(tasks):
    total_time = int(sum(float(task['time']) for task in tasks))
    total_hours = total_time // 60
    total_minutes = total_time % 60

    show_total_time = "{} mimutes".format(total_minutes)
    if total_hours:
        show_total_time = "{} heures et {}".format(total_hours, show_total_time)

    tasks = "\n".join(TASK_FORMAT.format(
        name=task['name'],
        time=float(task['time']),
        penibility=task['penibility'],
    ) for task in tasks)

    return "Temps total : {}\n\n{}".format(show_total_time, tasks)


# locale.setlocale(locale.LC_TIME, "fr_FR")
wb = load_workbook(WORBOOK_PATH)
ws = wb.active
tasks_per_week = []
day = FIRST_DAY
while day < LAST_DAY:
    day += ONE_WEEK
    week_number = get_week_number(day)
    week_details = {
        "name": NAME_STRING.format(
            first_day=serialize_date(day),
            last_day=serialize_date(day + ONE_WEEK - ONE_DAY, year=True),
        ),
        'tasks': defaultdict(list),
    }
    for row_number in range(FIRST_ROW, MAX_ROWS):
        row = get_row(ws, row_number)
        if not row or not should_do_task(
            row['year_frequency'], week_number, row['shift']
        ) or row["hide"]:
            continue

        for lookup, participant in PARTICIPANTS:
            if lookup == "x"
                week_details['tasks']["participant"].append(row)
            else: 
                print("Repos bien mérité")
    tasks_per_week.append(week_details)

for sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])

ws = wb.create_sheet(OUTPUT_SHEET)
cell = ws.cell(1, 1)
ws.row_dimensions[1].height = TOP_ROW_HEIGHT
ws.column_dimensions['A'].width = WEEK_COLUMN_WIDTH
for col, (_, participant) in enumerate(PARTICIPANTS, start=2):
    ws.column_dimensions[get_column_letter(col)].width = TASK_COLUMN_WIDTH
    cell = ws.cell(1, col)
    cell.value = TODO_PARTICIPANT.format(participant)
    cell.font = PARTICIPANT_FONT
    cell.border = OUTER_BORDER
    cell.alignment = CENTER_ALIGNMENT

# semaine, doit faire Maxime, doit Faire Morgane
for row_number, week_details in enumerate(tasks_per_week, start=OUTPUT_FIRST_ROW):
    cell = ws.cell(row_number, 1)
    cell.value = week_details["name"]
    cell.alignment = CENTER_ALIGNMENT
    cell.font = WEEK_FONT
    for col, (_, participant) in enumerate(PARTICIPANTS, start=2):
        cell = ws.cell(row_number, col)
        cell.value = format_tasks(week_details['tasks'][participant])
        cell.font = TASK_FONT
        cell.alignment = TASK_CELL_ALIGNMENT
    
    ws.row_dimensions[row_number].height = TASK_ROW_HEIGHT

wb.save('out.xlsx')
